#!/usr/bin/env python3
"""Convert a validated SOMA YAML file to a formatted Excel workbook.

Reads SOMA Container YAML data and produces an Excel workbook with tabs
for each entity type, styled consistently with the SOMA project conventions.

Usage:
    uv run python scripts/yaml_to_excel.py --input <yaml> --output <xlsx>
    uv run python scripts/yaml_to_excel.py --input <yaml> --output <xlsx> --template project/excel/soma.xlsx
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styling (matches generate_paper_excel.py)
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

# Tab colors by entity type
TAB_COLORS = {
    "Metadata": "4472C4",
    "Protocol": "70AD47",
    "ExposureCondition": "FFC000",
    "KeyEvent": "ED7D31",
    "CellularSystem": "4472C4",
    "InVivoSubject": "4472C4",
    "CFTRFunctionAssay": "C00000",
    "CFTRFunctionOutput": "C00000",
    "GeneExpressionAssay": "5B9BD5",
    "GeneExpressionOutput": "5B9BD5",
    "GobletCellAssay": "A9D18E",
    "GobletCellOutput": "A9D18E",
    "BALFSputumAssay": "7030A0",
    "BALFSputumOutput": "7030A0",
    "LungFunctionAssay": "002060",
    "LungFunctionOutput": "002060",
    "FoxJExpressionAssay": "BF8F00",
    "FoxJExpressionOutput": "BF8F00",
    "CiliaryFunctionAssay": "548235",
    "CiliaryFunctionOutput": "548235",
    "ASLAssay": "2E75B6",
    "ASLOutput": "2E75B6",
    "MucociliaryClearanceAssay": "843C0C",
    "MucociliaryClearanceOutput": "843C0C",
    "OxidativeStressAssay": "BF8F00",
    "OxidativeStressOutput": "BF8F00",
    "EGFRSignalingAssay": "ED7D31",
    "EGFRSignalingOutput": "ED7D31",
}

# ---------------------------------------------------------------------------
# Header definitions per tab (matching generate_paper_excel.py / scaffold)
# ---------------------------------------------------------------------------

HEADERS = {
    "Protocol": [
        "id", "name", "description", "protocol_type", "protocol_version",
        "equipment_required",
    ],
    "ExposureCondition": [
        "id", "name", "exposure_agent", "exposure_agent_id",
        "exposure_concentration_value", "exposure_concentration_unit",
        "exposure_duration_value", "exposure_duration_unit",
    ],
    "KeyEvent": [
        "id", "name", "description", "biological_action",
        "level_of_biological_organization",
    ],
    "CellularSystem": [
        "id", "name", "description", "subject_type",
        "cell_line", "cell_line_id", "primary_cell", "cell_type", "cell_type_id",
        "anatomical_origin", "model_species", "model_species_id",
        "cell_culture_growth_mode", "substrate_type",
        "days_at_differentiation", "donor_info",
    ],
    "InVivoSubject": [
        "id", "name", "description", "subject_type",
        "model_species", "model_species_id",
        "age_value", "age_unit", "sex",
        "subject_characteristics", "disease_state",
        "sample_type", "collection_site",
    ],
    "CFTRFunctionAssay": [
        "id", "name", "description",
        "stimulation_agent", "inhibitor_used",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "CFTRFunctionOutput": [
        "id", "name", "description",
        "cftr_chloride_secretion_value", "cftr_chloride_secretion_unit",
        "cftr_forskolin_response_value", "cftr_forskolin_response_unit",
        "inhibitor_sensitive_current_value", "inhibitor_sensitive_current_unit",
        "source_assay",
    ],
    "GeneExpressionAssay": [
        "id", "name", "description",
        "target_gene", "gene_expression_method", "normalization_reference",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "GeneExpressionOutput": [
        "id", "name", "description",
        "mrna_level_value", "mrna_level_unit",
        "protein_level_value", "protein_level_unit",
        "percentage_positive_cells_value", "percentage_positive_cells_unit",
        "source_assay",
    ],
    "GobletCellAssay": [
        "id", "name", "description",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "GobletCellOutput": [
        "id", "name", "description",
        "goblet_cell_percentage_value", "goblet_cell_percentage_unit",
        "muc5ac_mrna_expression_value", "muc5ac_mrna_expression_unit",
        "muc5ac_protein_expression_value", "muc5ac_protein_expression_unit",
        "muc5b_mrna_expression_value", "muc5b_mrna_expression_unit",
        "muc5b_protein_expression_value", "muc5b_protein_expression_unit",
        "mucin_secretion_rate_value", "mucin_secretion_rate_unit",
        "source_assay",
    ],
    "BALFSputumAssay": [
        "id", "name", "description",
        "target_cell_type",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "BALFSputumOutput": [
        "id", "name", "description",
        "il6_concentration_value", "il6_concentration_unit",
        "source_assay",
    ],
    "LungFunctionAssay": [
        "id", "name", "description",
        "reference_dataset",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "LungFunctionOutput": [
        "id", "name", "description",
        "lung_resistance_value", "lung_resistance_unit",
        "source_assay",
    ],
    "FoxJExpressionAssay": [
        "id", "name", "description",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "FoxJExpressionOutput": [
        "id", "name", "description",
        "foxj1_mrna_expression_value", "foxj1_mrna_expression_unit",
        "foxj1_positive_cell_percentage_value", "foxj1_positive_cell_percentage_unit",
        "source_assay",
    ],
    "CiliaryFunctionAssay": [
        "id", "name", "description",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "CiliaryFunctionOutput": [
        "id", "name", "description",
        "cbf_value", "cbf_unit",
        "source_assay",
    ],
    "ASLAssay": [
        "id", "name", "description",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "ASLOutput": [
        "id", "name", "description",
        "asl_height_value", "asl_height_unit",
        "source_assay",
    ],
    "MucociliaryClearanceAssay": [
        "id", "name", "description",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "MucociliaryClearanceOutput": [
        "id", "name", "description",
        "transport_rate_value", "transport_rate_unit",
        "source_assay",
    ],
    "OxidativeStressAssay": [
        "id", "name", "description",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "OxidativeStressOutput": [
        "id", "name", "description",
        "ros_level_value", "ros_level_unit",
        "source_assay",
    ],
    "EGFRSignalingAssay": [
        "id", "name", "description",
        "informs_on_key_event", "study_subject",
        "has_exposure_condition", "follows_protocols",
        "assay_date",
    ],
    "EGFRSignalingOutput": [
        "id", "name", "description",
        "egfr_phosphorylation_value", "egfr_phosphorylation_unit",
        "source_assay",
    ],
}

# Mapping from YAML collection key -> (assay tab name, output tab name)
COLLECTION_MAP = {
    "cftr_assays": ("CFTRFunctionAssay", "CFTRFunctionOutput"),
    "gene_expression_assays": ("GeneExpressionAssay", "GeneExpressionOutput"),
    "goblet_cell_assays": ("GobletCellAssay", "GobletCellOutput"),
    "balf_sputum_assays": ("BALFSputumAssay", "BALFSputumOutput"),
    "lung_function_assays": ("LungFunctionAssay", "LungFunctionOutput"),
    "foxj_assays": ("FoxJExpressionAssay", "FoxJExpressionOutput"),
    "ciliary_function_assays": ("CiliaryFunctionAssay", "CiliaryFunctionOutput"),
    "asl_assays": ("ASLAssay", "ASLOutput"),
    "mcc_assays": ("MucociliaryClearanceAssay", "MucociliaryClearanceOutput"),
    "oxidative_stress_assays": ("OxidativeStressAssay", "OxidativeStressOutput"),
    "egfr_signaling_assays": ("EGFRSignalingAssay", "EGFRSignalingOutput"),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def _add_row(ws, row, data):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def _auto_width(ws, max_col, min_w=12, max_w=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(
            max(min_w, max(len(str(c.value or "")) for c in ws[letter])),
            max_w,
        )


def _make_sheet(wb, name, headers, rows, tab_color=None):
    ws = wb.create_sheet(name)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    ncol = len(headers)
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    _style_header(ws, 1, ncol)
    for i, row_data in enumerate(rows, 2):
        _add_row(ws, i, row_data)
    _auto_width(ws, ncol)
    return ws


def _fmt_value_unit(obj):
    """Extract value and unit string from a measurement dict like {value: '0.5', unit: {id: ..., name: ...}}."""
    if not obj or not isinstance(obj, dict):
        return "", ""
    val = obj.get("value", "")
    unit_obj = obj.get("unit", {})
    if isinstance(unit_obj, dict):
        unit_name = unit_obj.get("name", "")
        unit_id = unit_obj.get("id", "")
        unit_str = f"{unit_name} ({unit_id})" if unit_id else unit_name
    else:
        unit_str = str(unit_obj)
    return str(val), unit_str


def _fmt_id_name(obj):
    """Format an object with id/name as 'name (id)' or just 'id'."""
    if not obj or not isinstance(obj, dict):
        return str(obj) if obj else ""
    obj_id = obj.get("id", "")
    obj_name = obj.get("name", "")
    if obj_name and obj_id:
        return f"{obj_name} ({obj_id})"
    return obj_id or obj_name or ""


def _fmt_list_refs(items):
    """Format a list of references as semicolon-separated id strings."""
    if not items:
        return ""
    if isinstance(items, dict):
        return items.get("id", "")
    if isinstance(items, list):
        return "; ".join(
            item.get("id", str(item)) if isinstance(item, dict) else str(item)
            for item in items
        )
    return str(items)


def _get_nested(d, *keys, default=""):
    """Safely traverse nested dicts."""
    current = d
    for key in keys:
        if isinstance(current, dict):
            current = current.get(key, default)
        else:
            return default
    return current if current is not None else default


# ---------------------------------------------------------------------------
# Metadata extraction from YAML header comments
# ---------------------------------------------------------------------------

def _extract_metadata(yaml_path):
    """Extract metadata from YAML file header comments."""
    rows = [["Field", "Value"]]
    with open(yaml_path) as f:
        in_header = False
        for line in f:
            line = line.rstrip()
            if line.startswith("# ===="):
                in_header = not in_header
                continue
            if line.startswith("# ") and not line.startswith("# ----"):
                text = line[2:].strip()
                if not text or text.startswith("==="):
                    continue
                if ":" in text and not text.startswith("http"):
                    key, _, val = text.partition(":")
                    rows.append([key.strip(), val.strip()])
                elif text.startswith("- "):
                    rows.append(["", text[2:].strip()])
                else:
                    rows.append(["Note", text])
    return rows


# ---------------------------------------------------------------------------
# Entity extraction
# ---------------------------------------------------------------------------

def _collect_exposure_conditions(data):
    """Collect unique exposure conditions from all assays."""
    seen = {}
    for coll_key in COLLECTION_MAP:
        for assay in data.get(coll_key, []) or []:
            conditions = assay.get("has_exposure_condition", [])
            if isinstance(conditions, dict):
                conditions = [conditions]
            for ec in conditions or []:
                ec_id = ec.get("id", "")
                if ec_id and ec_id not in seen:
                    seen[ec_id] = ec
    return list(seen.values())


def _collect_key_events(data):
    """Collect unique key events from all assays."""
    seen = {}
    for coll_key in COLLECTION_MAP:
        for assay in data.get(coll_key, []) or []:
            ke = assay.get("informs_on_key_event")
            if ke and isinstance(ke, dict):
                ke_id = ke.get("id", "")
                if ke_id and ke_id not in seen:
                    seen[ke_id] = ke
    return list(seen.values())


def _collect_subjects(data):
    """Collect unique study subjects, split into CellularSystem and InVivoSubject."""
    cellular = {}
    invivo = {}
    for coll_key in COLLECTION_MAP:
        for assay in data.get(coll_key, []) or []:
            subj = assay.get("study_subject")
            if not subj or not isinstance(subj, dict):
                continue
            subj_id = subj.get("id", "")
            subj_type = subj.get("subject_type", "")
            if subj_type == "InVivoSubject":
                if subj_id not in invivo:
                    invivo[subj_id] = subj
            else:
                if subj_id not in cellular:
                    cellular[subj_id] = subj
    return list(cellular.values()), list(invivo.values())


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------

def _protocol_row(p):
    equip = p.get("equipment_required", [])
    if isinstance(equip, list):
        equip = "; ".join(str(e) for e in equip)
    return (
        p.get("id", ""),
        p.get("name", ""),
        p.get("description", ""),
        p.get("protocol_type", ""),
        p.get("protocol_version", ""),
        equip,
    )


def _exposure_row(ec):
    agent = ec.get("exposure_agent", {})
    conc = ec.get("exposure_concentration", {})
    dur = ec.get("exposure_duration", {})
    agent_name = agent.get("name", "") if isinstance(agent, dict) else str(agent)
    agent_id = agent.get("id", "") if isinstance(agent, dict) else ""
    conc_val, conc_unit = _fmt_value_unit(conc)
    dur_val, dur_unit = _fmt_value_unit(dur)
    return (
        ec.get("id", ""),
        ec.get("name", ""),
        agent_name,
        agent_id,
        conc_val,
        conc_unit,
        dur_val,
        dur_unit,
    )


def _key_event_row(ke):
    return (
        ke.get("id", ""),
        ke.get("name", ""),
        ke.get("description", ""),
        ke.get("biological_action", ""),
        ke.get("level_of_biological_organization", ""),
    )


def _cellular_system_row(subj):
    ct = subj.get("cell_type", {})
    cl = subj.get("cell_line", {})
    pc = subj.get("primary_cell", {})
    ao = subj.get("anatomical_origin", {})
    sp = subj.get("model_species", {})
    return (
        subj.get("id", ""),
        subj.get("name", ""),
        subj.get("description", ""),
        subj.get("subject_type", "CellularSystem"),
        _get_nested(cl, "name"),
        _get_nested(cl, "id"),
        _get_nested(pc, "name"),
        _get_nested(ct, "name"),
        _get_nested(ct, "id"),
        _fmt_id_name(ao),
        _get_nested(sp, "name"),
        _get_nested(sp, "id"),
        subj.get("cell_culture_growth_mode", ""),
        subj.get("substrate_type", ""),
        subj.get("days_at_differentiation", ""),
        subj.get("donor_info", ""),
    )


def _invivo_subject_row(subj):
    sp = subj.get("model_species", {})
    age = subj.get("age", {})
    age_val, age_unit = _fmt_value_unit(age)
    return (
        subj.get("id", ""),
        subj.get("name", ""),
        subj.get("description", ""),
        subj.get("subject_type", "InVivoSubject"),
        _get_nested(sp, "name"),
        _get_nested(sp, "id"),
        age_val,
        age_unit,
        subj.get("sex", ""),
        subj.get("subject_characteristics", ""),
        subj.get("disease_state", ""),
        subj.get("sample_type", ""),
        subj.get("collection_site", ""),
    )


def _assay_row(assay, assay_headers):
    """Build a generic assay row based on headers."""
    row = []
    for h in assay_headers:
        if h == "informs_on_key_event":
            row.append(_get_nested(assay, "informs_on_key_event", "id"))
        elif h == "study_subject":
            row.append(_get_nested(assay, "study_subject", "id"))
        elif h == "has_exposure_condition":
            conditions = assay.get("has_exposure_condition", [])
            if isinstance(conditions, dict):
                conditions = [conditions]
            row.append(_fmt_list_refs(conditions))
        elif h == "follows_protocols":
            protos = assay.get("follows_protocols", [])
            row.append(_fmt_list_refs(protos))
        else:
            row.append(assay.get(h, ""))
    return tuple(row)


def _output_row(output, output_headers):
    """Build a generic output row based on headers."""
    row = []
    for h in output_headers:
        if h == "source_assay":
            row.append(output.get("source_assay", ""))
        elif h.endswith("_value"):
            # e.g., cftr_chloride_secretion_value -> cftr_chloride_secretion.value
            slot_name = h[:-6]  # remove '_value'
            row.append(_get_nested(output, slot_name, "value"))
        elif h.endswith("_unit"):
            slot_name = h[:-5]  # remove '_unit'
            unit_obj = _get_nested(output, slot_name, "unit")
            if isinstance(unit_obj, dict):
                name = unit_obj.get("name", "")
                uid = unit_obj.get("id", "")
                row.append(f"{name} ({uid})" if uid else name)
            else:
                row.append(str(unit_obj) if unit_obj else "")
        else:
            row.append(output.get(h, ""))
    return tuple(row)


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def yaml_to_excel(input_path, output_path, template_path=None):
    """Convert a SOMA YAML file to an Excel workbook."""
    input_path = Path(input_path)
    output_path = Path(output_path)

    # Load YAML
    with open(input_path) as f:
        data = yaml.safe_load(f)

    if not data:
        print(f"Error: {input_path} is empty or invalid YAML", file=sys.stderr)
        sys.exit(1)

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Metadata tab ---
    meta_rows = _extract_metadata(input_path)
    if len(meta_rows) > 1:  # more than just the header
        ws = wb.create_sheet("Metadata")
        ws.sheet_properties.tabColor = TAB_COLORS.get("Metadata", "4472C4")
        for i, row in enumerate(meta_rows, 1):
            for j, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.alignment = WRAP
                cell.border = THIN_BORDER
                if i == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 80

    # --- Protocol tab ---
    protocols = data.get("protocols", []) or []
    if protocols:
        _make_sheet(
            wb, "Protocol", HEADERS["Protocol"],
            [_protocol_row(p) for p in protocols],
            tab_color=TAB_COLORS.get("Protocol"),
        )

    # --- ExposureCondition tab ---
    exposures = _collect_exposure_conditions(data)
    if exposures:
        _make_sheet(
            wb, "ExposureCondition", HEADERS["ExposureCondition"],
            [_exposure_row(ec) for ec in exposures],
            tab_color=TAB_COLORS.get("ExposureCondition"),
        )

    # --- KeyEvent tab ---
    key_events = _collect_key_events(data)
    if key_events:
        _make_sheet(
            wb, "KeyEvent", HEADERS["KeyEvent"],
            [_key_event_row(ke) for ke in key_events],
            tab_color=TAB_COLORS.get("KeyEvent"),
        )

    # --- CellularSystem and InVivoSubject tabs ---
    cellular, invivo = _collect_subjects(data)
    if cellular:
        _make_sheet(
            wb, "CellularSystem", HEADERS["CellularSystem"],
            [_cellular_system_row(s) for s in cellular],
            tab_color=TAB_COLORS.get("CellularSystem"),
        )
    if invivo:
        _make_sheet(
            wb, "InVivoSubject", HEADERS["InVivoSubject"],
            [_invivo_subject_row(s) for s in invivo],
            tab_color=TAB_COLORS.get("InVivoSubject"),
        )

    # --- Assay + Output tabs ---
    for coll_key, (assay_tab, output_tab) in COLLECTION_MAP.items():
        assays = data.get(coll_key, []) or []
        if not assays:
            continue

        assay_headers = HEADERS.get(assay_tab, [])
        output_headers = HEADERS.get(output_tab, [])

        # Build assay rows
        assay_rows = [_assay_row(a, assay_headers) for a in assays]
        _make_sheet(
            wb, assay_tab, assay_headers, assay_rows,
            tab_color=TAB_COLORS.get(assay_tab),
        )

        # Build output rows from has_specified_output
        output_rows = []
        for a in assays:
            out = a.get("has_specified_output")
            if out and isinstance(out, dict):
                # Add source_assay reference
                out_with_ref = dict(out)
                if "source_assay" not in out_with_ref:
                    out_with_ref["source_assay"] = a.get("id", "")
                output_rows.append(_output_row(out_with_ref, output_headers))

        if output_rows:
            _make_sheet(
                wb, output_tab, output_headers, output_rows,
                tab_color=TAB_COLORS.get(output_tab),
            )

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Convert SOMA YAML to formatted Excel workbook"
    )
    parser.add_argument(
        "--input", "-i", required=True,
        help="Path to SOMA Container YAML file",
    )
    parser.add_argument(
        "--output", "-o", required=True,
        help="Path for output Excel file",
    )
    parser.add_argument(
        "--template", "-t", default=None,
        help="Path to LinkML-generated Excel scaffold (default: project/excel/soma.xlsx)",
    )
    args = parser.parse_args()
    yaml_to_excel(args.input, args.output, args.template)


if __name__ == "__main__":
    main()
