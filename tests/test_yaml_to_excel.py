"""Tests for yaml_to_excel.py converter."""
import os
import subprocess
import tempfile

import pytest

SCRIPT = os.path.join(os.path.dirname(__file__), "..", "scripts", "yaml_to_excel.py")
DATA_DIR = os.path.join(os.path.dirname(__file__), "data", "valid")

EXAMPLE_YAMLS = [
    (
        "Container-liu2024-pm25-cftr.yaml",
        {
            "required_tabs": [
                "Metadata", "Protocol", "ExposureCondition", "KeyEvent",
                "CellularSystem", "InVivoSubject",
                "CFTRFunctionAssay", "CFTRFunctionOutput",
                "GeneExpressionAssay", "GeneExpressionOutput",
                "GobletCellAssay", "GobletCellOutput",
                "BALFSputumAssay", "BALFSputumOutput",
                "LungFunctionAssay", "LungFunctionOutput",
            ],
            "min_assay_rows": {
                "CFTRFunctionAssay": 4,
                "GeneExpressionAssay": 5,
                "GobletCellAssay": 2,
                "LungFunctionAssay": 2,
            },
        },
    ),
    (
        "Container-montgomery2020-pm25-mucociliary.yaml",
        {
            "required_tabs": [
                "Metadata", "Protocol", "ExposureCondition", "KeyEvent",
                "CellularSystem",
                "GeneExpressionAssay", "GeneExpressionOutput",
                "GobletCellAssay", "GobletCellOutput",
                "FoxJExpressionAssay", "FoxJExpressionOutput",
            ],
            "min_assay_rows": {
                "GeneExpressionAssay": 6,
                "GobletCellAssay": 4,
                "FoxJExpressionAssay": 2,
            },
        },
    ),
]


@pytest.mark.parametrize("yaml_name,expectations", EXAMPLE_YAMLS, ids=[e[0] for e in EXAMPLE_YAMLS])
def test_yaml_to_excel(yaml_name, expectations):
    """Test that yaml_to_excel.py produces a valid Excel file with expected tabs and rows."""
    import openpyxl

    yaml_path = os.path.join(DATA_DIR, yaml_name)
    assert os.path.exists(yaml_path), f"YAML file not found: {yaml_path}"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        output_path = tmp.name

    try:
        result = subprocess.run(
            ["uv", "run", "python", SCRIPT, "--input", yaml_path, "--output", output_path],
            capture_output=True, text=True, timeout=60,
        )
        assert result.returncode == 0, f"Script failed:\nstdout: {result.stdout}\nstderr: {result.stderr}"
        assert os.path.exists(output_path), "Output file was not created"

        wb = openpyxl.load_workbook(output_path)
        sheet_names = wb.sheetnames

        # Check required tabs exist
        for tab in expectations["required_tabs"]:
            assert tab in sheet_names, f"Missing tab: {tab}. Found: {sheet_names}"

        # Check minimum row counts (header row + data rows)
        for tab, min_rows in expectations.get("min_assay_rows", {}).items():
            ws = wb[tab]
            data_rows = ws.max_row - 1  # subtract header
            assert data_rows >= min_rows, (
                f"Tab '{tab}' has {data_rows} data rows, expected >= {min_rows}"
            )

        # Verify no empty tabs for collections that should have data
        for tab in expectations["required_tabs"]:
            ws = wb[tab]
            assert ws.max_row >= 2, f"Tab '{tab}' appears empty (only header row)"

        wb.close()
    finally:
        if os.path.exists(output_path):
            os.unlink(output_path)
